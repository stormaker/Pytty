import os
import pandas as pd
from openpyxl import Workbook, load_workbook

# 初始化一个新的Excel工作簿
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Aggregated Data"

# 为新工作簿设置列标题（如果需要）
new_ws['A1'] = 'File Name'
new_ws['B1'] = '订单号'
new_ws['C1'] = '订购日期'
new_ws['D1'] = '品名'
new_ws['E1'] = '规格'
new_ws['F1'] = '品名'
new_ws['G1'] = '成分'
new_ws['H1'] = '颜色'
new_ws['I1'] = '色号'
new_ws['J1'] = '数量'
new_ws['K1'] = '单价'
new_ws['L1'] = '备注'
new_ws['M1'] = '2nd Item'

# 初始行号
row_number = 2

# 遍历包含Excel文件的目录
for filename in os.listdir('.'):
    if filename.endswith('.xlsx') and filename != 'new_file.xlsx':  # 避免读取新文件（如果它已经存在）
        print(f"Reading {filename}")

        # 加载单个Excel文件
        wb = load_workbook(filename)
        ws = wb.active  # 默认读取第一个工作表

        # 从A1, A2, A3读取数据
        order_data = ws['B2'].value
        date_data = ws['F2'].value
        item_data = ws['A10'].value
        spec_data = ws['C10'].value
        component_data = ws['D10'].value
        color_data = ws['E10'].value
        color_code_data = ws['F10'].value
        qty_data = ws['G10'].value
        price_data = ws['H10'].value
        remark_date = ws['A14'].value
        second_item_date = ws['A11'].value

        # 将数据写入新的Excel文件
        new_ws[f'A{row_number}'] = filename
        new_ws[f'B{row_number}'] = order_data
        new_ws[f'C{row_number}'] = date_data
        new_ws[f'D{row_number}'] = item_data
        new_ws[f'E{row_number}'] = spec_data
        new_ws[f'F{row_number}'] = component_data
        new_ws[f'G{row_number}'] = color_data
        new_ws[f'H{row_number}'] = color_code_data
        new_ws[f'I{row_number}'] = qty_data
        new_ws[f'K{row_number}'] = price_data
        new_ws[f'L{row_number}'] = remark_date
        new_ws[f'M{row_number}'] = second_item_date
        # 更新行号
        row_number += 1

# 保存新的Excel文件
new_wb.save('new_file.xlsx')
print("Aggregated data saved to 'new_file.xlsx'")
